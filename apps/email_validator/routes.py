# -*- coding: utf-8 -*-
from flask import Blueprint, render_template, request, jsonify, session, redirect, url_for, flash, send_file
from models import db, User, Permission, App, EmailValidation, EmailResult
from functools import wraps
from datetime import datetime
from .validator import validate_email
import io
import csv
import json
from werkzeug.utils import secure_filename
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

email_validator_bp = Blueprint('email_validator', __name__)

def app_permission_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            flash('Por favor, faça login.', 'warning')
            return redirect(url_for('auth.login'))
        
        user = db.session.get(User, session['user_id'])
        if not user:
            session.clear()
            flash('Sessão inválida.', 'warning')
            return redirect(url_for('auth.login'))
        
        app = App.query.filter_by(route='/apps/email-validator').first()
        
        if not app:
            flash('App não encontrada.', 'danger')
            return redirect(url_for('dashboard'))
        
        if user.role != 'admin' and not user.has_permission(app.id):
            flash('Não tem permissão para aceder a esta app.', 'danger')
            return redirect(url_for('dashboard'))
        
        return f(*args, **kwargs)
    return decorated_function

# ==========================================
# ROTAS - CÓDIGO EXATO DO ORIGINAL
# ==========================================

@email_validator_bp.route('/')
@app_permission_required
def index():
    """Página principal - EXATO DO ORIGINAL"""
    return render_template('apps/email_validator.html')

@email_validator_bp.route('/api/validate', methods=['POST'])
@app_permission_required
def api_validate():
    """Validar email único - EXATO DO ORIGINAL"""
    data = request.get_json()
    email = data.get('email', '').strip()
    check_mx = data.get('checkMX', True)
    
    if not email:
        return jsonify({'error': 'Email não fornecido'}), 400
    
    # Validar email
    is_valid, reason = validate_email(email, check_mx)
    
    # Criar upload (sessão)
    user_id = session['user_id']
    validation = EmailValidation(user_id=user_id)
    db.session.add(validation)
    db.session.flush()
    
    # Guardar resultado
    email_result = EmailResult(
        validation_id=validation.id,
        email=email,
        is_valid=is_valid,
        is_duplicate=False,
        score=0,
        reason=reason,
        validation_type='single'
    )
    db.session.add(email_result)
    db.session.commit()
    
    return jsonify({
        'valid': is_valid,
        'reason': reason,
        'upload_id': validation.id
    })

@email_validator_bp.route('/api/upload', methods=['POST'])
@app_permission_required
def api_upload():
    """Upload de ficheiro - EXATO DO ORIGINAL"""
    if 'file' not in request.files:
        return jsonify({'error': 'Nenhum ficheiro'}), 400
    
    file = request.files['file']
    check_mx = request.form.get('checkMX', 'true') == 'true'
    
    if not file.filename:
        return jsonify({'error': 'Ficheiro vazio'}), 400
    
    filename = secure_filename(file.filename)
    file_ext = filename.rsplit('.', 1)[1].lower() if '.' in filename else ''
    
    # Ler emails do ficheiro
    emails = []
    
    try:
        if file_ext == 'csv':
            content = file.stream.read().decode('utf-8-sig')
            reader = csv.reader(io.StringIO(content))
            for row in reader:
                if row and row[0].strip():
                    emails.append(row[0].strip())
        
        elif file_ext == 'txt':
            content = file.stream.read().decode('utf-8-sig')
            emails = [line.strip() for line in content.split('\n') if line.strip()]
        
        elif file_ext in ['xlsx', 'xls']:
            workbook = openpyxl.load_workbook(file)
            sheet = workbook.active
            for row in sheet.iter_rows(values_only=True):
                if row and row[0]:
                    emails.append(str(row[0]).strip())
        else:
            return jsonify({'error': 'Formato não suportado'}), 400
    
    except Exception as e:
        return jsonify({'error': f'Erro ao ler ficheiro: {str(e)}'}), 400
    
    if not emails:
        return jsonify({'error': 'Nenhum email encontrado'}), 400
    
    # Criar upload (sessão)
    user_id = session['user_id']
    validation = EmailValidation(user_id=user_id)
    db.session.add(validation)
    db.session.flush()
    
    # Processar emails - MANTÉM DUPLICADOS como original
    seen_emails = {}
    
    for email in emails:
        email_lower = email.lower()
        
        # Verificar se é duplicado
        if email_lower in seen_emails:
            # Duplicado
            email_result = EmailResult(
                validation_id=validation.id,
                email=email,
                is_valid=False,
                is_duplicate=True,
                score=0,
                reason='Duplicado',
                validation_type='bulk'
            )
        else:
            # Validar
            is_valid, reason = validate_email(email, check_mx)
            email_result = EmailResult(
                validation_id=validation.id,
                email=email,
                is_valid=is_valid,
                is_duplicate=False,
                score=0,
                reason=reason,
                validation_type='bulk'
            )
            seen_emails[email_lower] = True
        
        db.session.add(email_result)
    
    db.session.commit()
    
    return jsonify({
        'success': True,
        'upload_id': validation.id
    })

@email_validator_bp.route('/api/history')
@app_permission_required
def api_history():
    """Histórico - EXATO DO ORIGINAL"""
    user_id = session['user_id']
    
    validations = EmailValidation.query.filter_by(user_id=user_id)\
        .order_by(EmailValidation.upload_date.desc())\
        .all()
    
    history = []
    for v in validations:
        history.append({
            'id': v.id,
            'upload_date': v.upload_date.strftime('%Y-%m-%d %H:%M:%S'),
            'total': v.total_emails,
            'valid': v.valid_count,
            'invalid': v.invalid_count
        })
    
    return jsonify(history)

@email_validator_bp.route('/api/upload/<int:upload_id>')
@app_permission_required
def api_upload_details(upload_id):
    """Detalhes de upload - EXATO DO ORIGINAL"""
    user_id = session['user_id']
    
    validation = db.session.get(EmailValidation, upload_id)
    
    if not validation or validation.user_id != user_id:
        return jsonify({'error': 'Não encontrado'}), 404
    
    # Buscar emails com paginação
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 30, type=int)
    filter_type = request.args.get('filter', 'all')
    
    query = EmailResult.query.filter_by(validation_id=upload_id)
    
    # Filtros
    if filter_type == 'valid':
        query = query.filter_by(is_valid=True, is_duplicate=False)
    elif filter_type == 'invalid':
        query = query.filter_by(is_valid=False, is_duplicate=False)
    
    query = query.order_by(EmailResult.id)
    
    # Paginação
    total = query.count()
    emails = query.limit(per_page).offset((page - 1) * per_page).all()
    
    results = []
    for e in emails:
        results.append({
            'id': e.id,
            'email': e.email,
            'valid': e.is_valid,
            'duplicate': e.is_duplicate,
            'reason': e.reason,
            'upload_date': e.upload_date.strftime('%Y-%m-%d %H:%M:%S')
        })
    
    return jsonify({
        'upload_id': validation.id,
        'upload_date': validation.upload_date.strftime('%Y-%m-%d %H:%M:%S'),
        'total': validation.total_emails,
        'valid': validation.valid_count,
        'invalid': validation.invalid_count,
        'emails': results,
        'pagination': {
            'page': page,
            'per_page': per_page,
            'total': total,
            'pages': (total + per_page - 1) // per_page
        }
    })

@email_validator_bp.route('/api/delete/<int:upload_id>', methods=['POST'])
@app_permission_required
def api_delete(upload_id):
    """Eliminar - EXATO DO ORIGINAL"""
    user_id = session['user_id']
    
    validation = db.session.get(EmailValidation, upload_id)
    
    if not validation or validation.user_id != user_id:
        return jsonify({'error': 'Não encontrado'}), 404
    
    db.session.delete(validation)
    db.session.commit()
    
    return jsonify({'success': True})

@email_validator_bp.route('/export/<int:upload_id>')
@app_permission_required
def export_excel(upload_id):
    """Export Excel - EXATO DO ORIGINAL"""
    user_id = session['user_id']
    
    validation = db.session.get(EmailValidation, upload_id)
    
    if not validation or validation.user_id != user_id:
        return jsonify({'error': 'Não encontrado'}), 404
    
    emails = EmailResult.query.filter_by(validation_id=upload_id)\
        .order_by(EmailResult.id).all()
    
    # Criar Excel
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = 'Emails'
    
    # Headers
    headers = ['ID', 'DATA DE UPLOAD', 'EMAIL', 'TIPO', 'RAZÃO']
    for col, header in enumerate(headers, 1):
        cell = sheet.cell(1, col, header)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='1F4788', end_color='1F4788', fill_type='solid')
        cell.alignment = Alignment(horizontal='center')
    
    # Data
    for row_idx, e in enumerate(emails, 2):
        sheet.cell(row_idx, 1, row_idx - 1)
        sheet.cell(row_idx, 2, e.upload_date.strftime('%Y-%m-%d %H:%M:%S'))
        sheet.cell(row_idx, 3, e.email)
        
        if e.is_duplicate:
            tipo = 'DUPLICADO'
            fill_color = 'FFC7CE'
        elif e.is_valid:
            tipo = 'VÁLIDO'
            fill_color = 'C6EFCE'
        else:
            tipo = 'INVÁLIDO'
            fill_color = 'FFC7CE'
        
        tipo_cell = sheet.cell(row_idx, 4, tipo)
        tipo_cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
        
        sheet.cell(row_idx, 5, e.reason)
    
    # Larguras
    sheet.column_dimensions['A'].width = 8
    sheet.column_dimensions['B'].width = 20
    sheet.column_dimensions['C'].width = 35
    sheet.column_dimensions['D'].width = 15
    sheet.column_dimensions['E'].width = 40
    
    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'emails_{timestamp}.xlsx'
    )

@email_validator_bp.route('/details')
@app_permission_required
def details():
    """Página de detalhes - EXATO DO ORIGINAL"""
    upload_id = request.args.get('id', type=int)
    
    if not upload_id:
        flash('ID inválido', 'danger')
        return redirect(url_for('email_validator.index'))
    
    user_id = session['user_id']
    validation = db.session.get(EmailValidation, upload_id)
    
    if not validation or validation.user_id != user_id:
        flash('Validação não encontrada', 'danger')
        return redirect(url_for('email_validator.index'))
    
    return render_template('apps/email_validator_details.html', upload_id=upload_id)